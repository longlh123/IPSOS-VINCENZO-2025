import sys
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
import numpy as np

import logging
from PyQt5.QtWidgets import (
    QMainWindow, QWidget, QVBoxLayout, QPushButton, QLabel,
    QFileDialog, QMessageBox, QAction, QGroupBox, QGridLayout, QComboBox, QHBoxLayout, QScrollArea, QFrame, QSizePolicy, QButtonGroup, QRadioButton
)
from PyQt5.Qt import Qt
from ui.widgets.multi_select import MultiSelectWidget
from ui.widgets.ReportChartWidget import ReportChartWidget

from models.chart_mapping import map_chart_data, map_calculation_chart_components

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import (
    get_column_letter,
    column_index_from_string
)


class ReportTab(QWidget):

    def __init__(self, data=None, dataset=None):
        super().__init__()
        
        #Set up loggin
        self.logger = logging.getLogger(__name__)

        self.data = data.copy() if data is not None else pd.DataFrame()
        
        self.dataset = dataset.copy()
        
        #Create chart data
        self.chart_data = map_chart_data(self.data, self.dataset)

        self.exported_data = {}

        self.time_mode = dataset.get('time-mode', {})
        self.filters = dataset.get('main-filters', {})
        
        #Lưu trữ filter multiselection
        self.multiselectitems = {}

        #Lưu trữ filter combobox
        self.comboboxitems = {}
        
        # Main layout
        self.main_layout = QVBoxLayout(self)

        # Create Time Mode group
        time_mode_group = self.create_time_mode_group(self.time_mode)

        self.main_layout.addWidget(time_mode_group)
 
        # Create filter group (fixed at top)
        filter_group = self.create_filter_group(self.filters)
        # filter_group.setFixedHeight(120)  # optional: adjust as needed
        
        self.main_layout.addWidget(filter_group)

        #Export Button

        button_layout = QHBoxLayout()

        export_chart_data_button = QPushButton("Export Chart Data")
        export_chart_data_button.clicked.connect(self.export_chart_data_excel)

        export_rawdata_button = QPushButton("Export CSV Data")
        export_rawdata_button.clicked.connect(self.export_rawdata)

        button_layout.addWidget(export_chart_data_button)
        button_layout.addWidget(export_rawdata_button)

        self.main_layout.addLayout(button_layout)

        # Create a scroll area
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setFrameShape(QFrame.NoFrame)
            
        # Create a widget for the scroll area
        scroll_content = QWidget()
        scroll_content.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)

        self.scroll_layout = QVBoxLayout(scroll_content)
        self.scroll_layout.setSpacing(15)
        
        self.set_filters(self.chart_data, self.filters)

        self.render_chart()

        self.scroll_layout.addStretch() 

        # Set the scroll area widget
        scroll_area.setWidget(scroll_content)
        
        # Add scroll area to main layout
        self.main_layout.addWidget(scroll_area)

    def reset_filters_and_chart(self):
        # 1. Xóa filter group cũ
        for i in reversed(range(self.main_layout.count())):
            widget = self.main_layout.itemAt(i).widget()
            if isinstance(widget, QGroupBox) and widget.title() == "Filter":
                self.main_layout.takeAt(i)
                widget.deleteLater()

        # 2. Clear dict filter
        self.multiselectitems.clear()
        self.comboboxitems.clear()

        # 3. Tạo lại filter group mới
        filter_group = self.create_filter_group(self.filters)

        self.main_layout.insertWidget(1, filter_group)   # chèn lại vào layout chính (sau TimeMode group)

        self.set_filters(self.chart_data, self.filters)

        # 4. Clear chart cũ
        self.clear_layout()

        # 5. Render lại chart
        self.render_chart()

    def load_chart_data(self):

        filtered_chart_data = self.chart_data.copy()
        filtered_chart_data.reset_index(inplace=True)
        
        for key, values in self.filters.items():
            selected_items = []

            if key in self.multiselectitems:
                multiselectitem = self.multiselectitems[key]
                selected_items = multiselectitem.get_selected_items()
            elif key in self.comboboxitems:
                selected_items = [self.comboboxitems[key].currentText()]

            if len(selected_items) > 0:
                filtered_chart_data = filtered_chart_data.loc[filtered_chart_data[key].isin(selected_items)]
        
        return filtered_chart_data

    def get_group_items(self):
        group_items = self.dataset.get('group-by', []).copy()

        for time_mode_item in self.dataset.get('time-mode', []):
            if not bool(time_mode_item.get('checked')):
                group_items.remove(time_mode_item.get('label'))

        return group_items

    def render_chart(self):
        
        group_items = self.get_group_items()

        self.clear_layout()

        for chart_config in self.dataset.get('charts', []):
            filtered_chart_data = self.load_chart_data()

            calculated_chart_data = map_calculation_chart_components(filtered_chart_data, chart_config, group_items)
            
            chart_group = self.create_chart_group(calculated_chart_data, chart_config, group_items)
            self.scroll_layout.addWidget(chart_group)

            if chart_config['name'] not in self.exported_data.keys():
                self.exported_data[chart_config['name']] = pd.DataFrame() 
                
            self.exported_data[chart_config['name']] = calculated_chart_data
        
    def create_chart_group(self, chart_data, chart_config, group_by):
        chart_group = QGroupBox(chart_config.get('title'))
        chart_group.setMinimumHeight(700)

        layout = QVBoxLayout(chart_group)

        if "filter-mapping" in chart_config.keys():
            def on_filter_combobox_changed(selected):
                #clear_old_chart
                for i in reversed(range(layout.count())):
                    w = layout.itemAt(i).widget()
                    if isinstance(w, ReportChartWidget):
                        w.setParent(None)
                
                new_config = chart_config.copy()
                new_config['xAxis']['label'] = chart_config["filter-mapping"][selected]

                # tạo chart mới
                filtered_chart_data = self.load_chart_data()

                new_data = map_calculation_chart_components(filtered_chart_data, new_config, group_by)
                new_chart = ReportChartWidget(new_data, new_config, group_by)

                layout.addWidget(new_chart)

            filter_layout = QGridLayout()
            filter_layout.setColumnStretch(1, 1)
            filter_layout.setColumnStretch(3, 1)

            filter_layout.addWidget(QLabel("Product"), 0, 0)

            filter_box = QComboBox()
            filter_box.addItems(chart_config.get('filter-mapping', {}).keys())
            filter_box.currentTextChanged.connect(on_filter_combobox_changed)
            filter_box.setCurrentIndex(0)

            filter_layout.addWidget(filter_box, 0, 1)

            layout.addLayout(filter_layout)

        chart_widget = ReportChartWidget(chart_data, chart_config, group_by)

        layout.addWidget(chart_widget)

        return chart_group

    def create_time_mode_group(self, time_mode):
        groupbox = QGroupBox("TimeMode")

        layout = QGridLayout(groupbox)
        layout.setColumnStretch(1, 1)
        layout.setColumnStretch(3, 1)

        layout.addWidget(QLabel("Time Mode:"), 0, 0)

        self.time_mode_group = QButtonGroup(self)
        
        radio_layout = QHBoxLayout()

        for i, item in enumerate(self.time_mode):
            radio_item = QRadioButton(item.get('label'))
            self.time_mode_group.addButton(radio_item, i)

            radio_item.setChecked(item.get('checked', False))
            radio_item.toggled.connect(self.on_time_mode_changed)
            
            setattr(self, f"{item.get('name')}_radioitem", radio_item)
        
            radio_layout.addWidget(radio_item)
        
        layout.addLayout(radio_layout, 0, 1)

        return groupbox

    def on_time_mode_changed(self):
        checked_button = self.time_mode_group.checkedButton()

        if not checked_button:
            return
        
        mode = checked_button.text()

        for item in self.time_mode:
            item.update({'checked': int(item.get('label') == mode)})
        
        # Clear + tạo lại filter + chart
        self.reset_filters_and_chart()

    def create_filter_group(self, filters):
        groupbox = QGroupBox("Filter")
        
        layout = QGridLayout(groupbox)
        layout.setColumnStretch(1, 1)
        layout.setColumnStretch(3, 1)
        
        def create_filter_row(row, keys: list):
            row, col = row, 0

            for key in keys:
                label = QLabel(f"{key}")
                layout.addWidget(label, row, col)

                if isinstance(filters[key], str):
                    comboboxitem = QComboBox()
                     
                    comboboxitem.currentTextChanged.connect(
                        lambda text: self.handle_filter_changed(key, text)
                    )

                    layout.addWidget(comboboxitem, row, col + 1)

                    self.comboboxitems[key] = comboboxitem
                elif isinstance(filters[key], list):
                    multiselectitem = MultiSelectWidget(filters[key])

                    multiselectitem.selectionChanged.connect(
                        lambda items: self.handle_filter_changed(key, items)
                    )

                    layout.addWidget(multiselectitem, row, col + 1)

                    self.multiselectitems[key] = multiselectitem

                col = col + 2

        keys = list(filters.keys())

        for timemode_item in self.time_mode:
            if not bool(timemode_item.get('checked')):
                keys.remove(timemode_item.get('label'))
        
        for i in range(0, len(filters.keys()), 2):
            i_from = i 
            i_to = i + 2 if i < len(keys) else i + 1

            create_filter_row(i, keys[i_from:i_to])

        return groupbox

    def handle_filter_changed(self, column_name: str, value: str):

        # Cập nhật dữ liệu cho tab Report
        self.render_chart()

        print(f"{column_name} changed to {value}")
    
    def set_filters(self, data, filters):
        for column_name in filters.keys():
            filters[column_name] = data[column_name].dropna().unique().tolist()

            if column_name in self.multiselectitems:
                self.multiselectitems[column_name].set_items(filters[column_name])
            elif column_name in self.comboboxitems:
                self.comboboxitems[column_name].clear()
                self.comboboxitems[column_name].addItems(filters[column_name])
                self.comboboxitems[column_name].setCurrentIndex(0)

    def clear_layout(self):
        while self.scroll_layout.count():
            item = self.scroll_layout.takeAt(0)
            widget = item.widget()
            if widget is not None:
                widget.setParent(None)

    def export_rawdata(self):
        rawdata = self.data.copy()

        rawdata.reset_index(inplace=True)
        rawdata.rename(columns=self.dataset['renamed-columns'], inplace=True)
        
        for key, values in self.filters.items():
            selected_items = []

            if key in self.multiselectitems:
                multiselectitem = self.multiselectitems[key]
                selected_items = multiselectitem.get_selected_items()
            elif key in self.comboboxitems:
                selected_items = [self.comboboxitems[key].currentText()]

            if len(selected_items) > 0:
                rawdata = rawdata.loc[rawdata[key].isin(selected_items)]

        if rawdata is None or rawdata.empty:
            QMessageBox.warning(self, "Export Error", "No chart data to export.")
            return

        # mở hộp thoại chọn file save
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Export CSV",
            "output.csv",
            "CSV Files (*.csv)"
        )

        if file_path:
            if not file_path.endswith('.csv'):
                file_path += '.csv'

            try:
                rawdata.to_csv(file_path, index=False, encoding='utf-8-sig')
                QMessageBox.information(self, "Export Success", f"Chart data exported to:\n{file_path}")
            except Exception as ex:
                QMessageBox.critical(self, "Export Failed", f"An error occurred:\n{str(ex)}")

    def export_chart_data_excel(self):
        #Open file save dialog
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Export Excel",
            "output.xlsx",
            "Excel Files (*.xlsx)"
        )

        if file_path:
            if not file_path.endswith('.xlsx'):
                file_path += '.xlsx'

            try:
                self.logger.info(f"Exporting to: {file_path}")

                wb = openpyxl.Workbook()

                if "Sheet" in wb.sheetnames:
                    wb.remove(wb["Sheet"])
                
                for chart in self.dataset.get('charts', []):
                    chart_data = self.exported_data[chart['name']]
                    self.create_sheet(wb, chart_data=chart_data, sheet_name=chart['title'])

                wb.save(file_path)

                self.logger.info("Export Success.")
                QMessageBox.information(self, "Export Success", f"Chart data exported to:\n{file_path}")
            except Exception as ex:
                self.logger.error(f"An error occurred:\n{str(ex)}")
                QMessageBox.critical(self, "Export Failed", f"An error occurred:\n{str(ex)}")   

    def create_sheet(self, wb, chart_data=pd.DataFrame(), sheet_name=""):
        try:
            group_items = self.get_group_items()

            chart_data.set_index(group_items, inplace=True)

            sheet = wb.create_sheet(sheet_name)

            row = 2
            column = 2

            n_levels = chart_data.index.nlevels
            n_cols = len(chart_data.index)

            sheet.merge_cells(
                start_row = row,
                end_row = row + n_levels - 1,
                start_column = column - 1,
                end_column = column - 1 
            )

            for i in range(n_cols):
                group = chart_data.index[i]

                for j in range(n_levels):
                    cell = sheet.cell(row=row+j, column=column+i)

                    if j < n_levels - 1:
                        checked = False # Kiểm tra nếu dữ liệu đã được add hay chưa
                        
                        end_column = column + i #Tìm vị trí column cuối cùng để merge cell

                        for k in range(column + i):
                            prev_cell = sheet.cell(row=row+j, column=column+k)

                            if prev_cell.value == group[j]:
                                checked = True
                                start_column = column + k #Tìm vị trí column đầu tiên để merge cell
                                break
                        
                        if not checked:
                            cell.value = group[j]
                            cell.font = Font(bold=True)
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                            cell.border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin'),
                            )
                            end_column = column + i - 1

                            #Merge cell header
                            if i > 0:
                                sheet.merge_cells(
                                    start_row = row + j,
                                    end_row = row + j,
                                    start_column = start_column,
                                    end_column = end_column 
                                )
                        else:
                            if i == n_cols - 1:
                                sheet.merge_cells(
                                    start_row = row + j,
                                    end_row = row + j,
                                    start_column = start_column,
                                    end_column = end_column 
                                )
                    else:
                        cell.value = group if n_levels == 1 else group[j]
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'),
                        )

            row = 2 + n_levels

            for i, column_name in enumerate(list(chart_data.columns)):
                cell_axis = sheet.cell(row=row + i, column=1)

                cell_axis.value = column_name
                cell_axis.font = Font(bold=True)
                cell_axis.alignment = Alignment(horizontal='center', vertical='center')
                cell_axis.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'),
                )

                for j in range(n_cols):
                    group = chart_data.index[j]

                    cell_value = sheet.cell(row=row+i, column=column+j)

                    cell_value.value = chart_data.loc[group][column_name]
                    cell_value.number_format = "#,###.0"
                    cell_value.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin'),
                )
        except Exception as ex:
            raise(ex)
            

            





        

    

    

    
    
    